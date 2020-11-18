import openpyxl
import collections
import os
from PIL import Image
import random
import json
from shutil import copyfile

IMG_DIM = 224
xlsx_path = "../Annotations/kaggle-classes.xlsx" #"./Annotations/Mask-classes.xlsx"
photo_path_prefix = "../../combined" #"./Photos/"


multiclass_classifications = {
    "incorrect": 0,
    "correct": 1,
    "none": 2
}

def split_into_train_val_test(dict):
    random.seed(230)

    test = []
    val = []
    train = []

    for category in dict:
        img_names = list(dict[category])
        img_names.sort()
        random.shuffle(img_names)

        test_split = int(0.1 * len(img_names))
        val_split = int(.18 * len(img_names))

        test_img_names = img_names[:test_split]
        val_img_names = img_names[test_split: test_split + val_split]
        train_img_names = img_names[test_split + val_split:]

        test.extend(test_img_names)
        val.extend(val_img_names)
        train.extend(train_img_names)

    return {
        "test": set(test),
        "val": set(val),
        "train": set(train)
    }


def generate_binary_and_multiclass_dict():
    wb_obj = openpyxl.load_workbook(xlsx_path)
    sheet_obj = wb_obj.active

    multiclass_dict = collections.defaultdict(set)

    num_row = sheet_obj.max_row

    for i in range(2, num_row + 1):
        image_name = sheet_obj.cell(row=i, column=1).value
        binary = sheet_obj.cell(row=i, column=2).value
        multiclass = sheet_obj.cell(row=i, column=3).value

        multiclass_dict[multiclass].add(image_name)

    return multiclass_dict


def generate_train_val_test_split(multiclass_dict):
    multiclass_split = split_into_train_val_test(multiclass_dict)
    return multiclass_split


def get_split(img, split_dict):
    if img in split_dict["test"]:
        return "test"
    if img in split_dict["val"]:
        return "val"
    return "train"


def resize_and_save(filename, output_path, size=IMG_DIM):
    """Resize the image contained in `filename` and save it to the `output_dir`"""
    image = Image.open(filename)
    # Use bilinear interpolation instead of the default "nearest neighbor" method
    image = image.resize((size, size), Image.BILINEAR)
    image.save(output_path)


def copy_photo_files_into_directories(classification_dict, classification_type, split_dict):
    for (category, images) in classification_dict.items():
        num = str(classification_type[category])
        for img in images:
            split = get_split(img, split_dict)
            make_dir(os.path.join("three_classes", "multiclass", split))
            new_img_path = os.path.join("three_classes", "multiclass", split, num + "_" + img)
            resize_and_save(os.path.join(photo_path_prefix, img), new_img_path)


def categorize_train_val_test_split(verbose = False):
    multiclass_dict = generate_binary_and_multiclass_dict()
    if verbose:
        print("Finished categorizing pictures into their respective classes for multiclass classification")
    multiclass_split = generate_train_val_test_split(multiclass_dict)
    if verbose:
        print("Finished splitting dataset")
    copy_photo_files_into_directories(multiclass_dict, multiclass_classifications, multiclass_split)
    if verbose:
        print("Finished copying photos into the 'multiclass' folder")


def make_dir(path):
    path = os.path.abspath(os.path.join(path))

    if not os.path.exists(path):
        try:
            os.makedirs(path)
        except Exception as e:
            # Raise if directory can't be made, because image cuts won't be saved.
            print('Error creating directory')
            raise e

def main():
    categorize_train_val_test_split(True)


if __name__ == "__main__":
    main()